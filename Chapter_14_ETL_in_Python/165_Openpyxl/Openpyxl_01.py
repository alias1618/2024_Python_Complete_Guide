from openpyxl import load_workbook

wb = load_workbook("Chapter_14_ETL_in_Python/165_Openpyxl/14-3Dodgers.xlsx")
result = []

ws = wb.worksheets[0]
for row in ws.iter_rows():
    # List comprehension
    result.append([cell.value for cell in row])

print(result)